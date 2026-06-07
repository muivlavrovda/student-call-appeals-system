from datetime import timedelta
from io import BytesIO
from unittest.mock import patch

import pytest
from django.contrib.auth.models import Group
from django.urls import reverse
from django.utils import timezone
from django.utils.translation import gettext as _
from openpyxl import load_workbook

from appeals.classifier import ClassifyResult, ClassifyStatus
from appeals.models import Appeal, AppealComment, AppealHistoryEvent
from appeals.roles import ADMIN_GROUP, OPERATOR_GROUP, RESPONSIBLE_GROUP, sync_access_groups
from appeals.tests.factories import (
    AppealCategoryFactory,
    AppealCommentFactory,
    AppealFactory,
    AppealHistoryEventFactory,
    DepartmentFactory,
)
from users.tests.factories import UserFactory

LIST_URL_NAME = "appeals:appeal_list"
CREATE_URL_NAME = "appeals:appeal_create"
DETAIL_URL_NAME = "appeals:appeal_detail"
COMMENT_URL_NAME = "appeals:appeal_comment_create"
START_URL_NAME = "appeals:appeal_start_processing"
CLOSE_URL_NAME = "appeals:appeal_close"
TRANSFER_URL_NAME = "appeals:appeal_transfer"
REPORT_URL_NAME = "appeals:appeal_report"
REPORT_XLSX_URL_NAME = "appeals:appeal_report_xlsx"
REPORT_DOCX_URL_NAME = "appeals:appeal_report_docx"
REPORT_FILE_URL_NAME = "appeals:appeal_report_file"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DOCX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
LOGIN_URL = "/accounts/login/"


@pytest.fixture(autouse=True)
def _isolated_media(settings, tmp_path):
    # Выгрузка отчётов сохраняет копии на диск; направляем MEDIA_ROOT во
    # временный каталог, чтобы тесты не писали в реальную папку проекта.
    settings.MEDIA_ROOT = tmp_path


def _responsible_with_department(*, password: str = "secret"):
    responsible = _user_in_group(RESPONSIBLE_GROUP, password=password)
    department = DepartmentFactory()
    department.members.add(responsible)
    return responsible, department


def _valid_appeal_payload(category):
    return {
        "student_full_name": "Иванов Иван Иванович",
        "student_phone": "+7 (900) 123-45-67",
        "summary": "Вопрос по справке",
        "category": category.pk,
        "department": "",
        "description": "Студент просит справку об обучении.",
    }


def _user_in_group(*group_names: str, password: str = "secret"):
    sync_access_groups()
    user = UserFactory(password=password)
    user.groups.add(*Group.objects.filter(name__in=group_names))
    return user


# --- список обращений: контроль доступа --------------------------------------


@pytest.mark.django_db
@pytest.mark.security
def test_appeal_list_redirects_anonymous_to_login(client):
    response = client.get(reverse(LIST_URL_NAME))

    assert response.status_code == 302
    assert LOGIN_URL in response["Location"]


@pytest.mark.django_db
@pytest.mark.security
def test_appeal_list_forbidden_without_view_permission(client):
    user = UserFactory(password="secret")
    client.login(email=user.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME))

    assert response.status_code == 403


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_appeal_list_scopes_to_operator_own_appeals(client):
    operator = _user_in_group(OPERATOR_GROUP)
    own = AppealFactory(created_by=operator, summary="моя заявка")
    other = AppealFactory(summary="чужая заявка")
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME))
    appeals = list(response.context["appeals"])

    assert appeals == [own]
    assert "моя заявка" in response.content.decode()
    assert "чужая заявка" not in response.content.decode()
    assert other not in appeals


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_appeal_list_scopes_to_responsible_department(client):
    responsible = _user_in_group(RESPONSIBLE_GROUP)
    department = DepartmentFactory()
    department.members.add(responsible)
    in_scope = AppealFactory(department=department)
    out_of_scope = AppealFactory()
    client.login(email=responsible.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME))
    appeals = list(response.context["appeals"])

    assert appeals == [in_scope]
    assert out_of_scope not in appeals


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_appeal_list_shows_all_appeals_to_admin(client):
    admin = _user_in_group(ADMIN_GROUP)
    appeals = [AppealFactory(), AppealFactory()]
    client.login(email=admin.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME))

    assert set(response.context["appeals"]) == set(appeals)


# --- список обращений: отрисовка ---------------------------------------------


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_shows_empty_state(client):
    operator = _user_in_group(OPERATOR_GROUP)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME))

    assert response.status_code == 200
    assert "Обращений пока нет." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_marks_overdue_appeal(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory(
        created_by=operator,
        status=Appeal.Status.NEW,
        due_at=timezone.now() - timedelta(days=1),
    )
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME))

    assert "Просрочено" in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_is_paginated(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory.create_batch(25, created_by=operator)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME))

    assert response.context["is_paginated"] is True
    assert len(response.context["appeals"]) == 20


# --- список обращений: поиск, фильтрация, сортировка -------------------------


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_search_matches_name_and_summary(client):
    operator = _user_in_group(OPERATOR_GROUP)
    target = AppealFactory(
        created_by=operator,
        student_full_name="Петров Пётр",
        summary="Вопрос про общежитие",
    )
    AppealFactory(created_by=operator, student_full_name="Сидоров Иван", summary="Справка")
    client.login(email=operator.email, password="secret")

    by_name = client.get(reverse(LIST_URL_NAME), {"q": "Петров"})
    by_summary = client.get(reverse(LIST_URL_NAME), {"q": "общежитие"})

    assert list(by_name.context["appeals"]) == [target]
    assert list(by_summary.context["appeals"]) == [target]


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_search_matches_normalized_phone(client):
    operator = _user_in_group(OPERATOR_GROUP)
    # В базе телефон хранится нормализованным (7XXXXXXXXXX); ищем в «человеческом»
    # формате и всё равно находим.
    target = AppealFactory(created_by=operator, student_phone="+7 (900) 111-22-33")
    AppealFactory(created_by=operator, student_phone="+7 (900) 999-88-77")
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME), {"q": "8 (900) 111-22-33"})

    assert list(response.context["appeals"]) == [target]


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_filters_by_status(client):
    operator = _user_in_group(OPERATOR_GROUP)
    new_one = AppealFactory(created_by=operator, status=Appeal.Status.NEW)
    AppealFactory(created_by=operator, status=Appeal.Status.CLOSED)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME), {"status": Appeal.Status.NEW})

    assert list(response.context["appeals"]) == [new_one]


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_filters_by_category_and_department(client):
    admin = _user_in_group(ADMIN_GROUP)
    learning = DepartmentFactory(name="Учебный отдел")
    hr = DepartmentFactory(name="Отдел кадров")
    certificates = AppealCategoryFactory(name="Справки", department=learning)
    staffing = AppealCategoryFactory(name="Кадры", department=hr)
    cert_appeal = AppealFactory(category=certificates, department=learning)
    AppealFactory(category=staffing, department=hr)
    client.login(email=admin.email, password="secret")

    by_category = client.get(reverse(LIST_URL_NAME), {"category": certificates.pk})
    by_department = client.get(reverse(LIST_URL_NAME), {"department": learning.pk})

    assert list(by_category.context["appeals"]) == [cert_appeal]
    assert list(by_department.context["appeals"]) == [cert_appeal]


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_sorts_by_due_date(client):
    operator = _user_in_group(OPERATOR_GROUP)
    now = timezone.now()
    later = AppealFactory(created_by=operator, due_at=now + timedelta(days=10))
    sooner = AppealFactory(created_by=operator, due_at=now + timedelta(days=1))
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME), {"sort": "due_at"})

    assert list(response.context["appeals"]) == [sooner, later]


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_default_sort_is_newest_first(client):
    operator = _user_in_group(OPERATOR_GROUP)
    older = AppealFactory(created_by=operator, created_at=timezone.now() - timedelta(days=2))
    newer = AppealFactory(created_by=operator, created_at=timezone.now())
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME))

    assert list(response.context["appeals"]) == [newer, older]


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_ignores_unknown_sort(client):
    operator = _user_in_group(OPERATOR_GROUP)
    older = AppealFactory(created_by=operator, created_at=timezone.now() - timedelta(days=2))
    newer = AppealFactory(created_by=operator, created_at=timezone.now())
    client.login(email=operator.email, password="secret")

    # Неизвестное значение сортировки не должно влиять на запрос — откат к
    # сортировке по умолчанию (сначала новые).
    response = client.get(reverse(LIST_URL_NAME), {"sort": "student_phone"})

    assert list(response.context["appeals"]) == [newer, older]


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_shows_not_found_state_when_filtered(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory(created_by=operator, summary="Справка об обучении")
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME), {"q": "несуществующий запрос"})

    assert response.context["filters_active"] is True
    assert "не найдено" in response.content.decode()


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_appeal_list_filters_cannot_escape_access_scope(client):
    # Фильтр применяется поверх доступных заявок, поэтому подбор параметров под
    # чужую заявку (её отдел и статус) не выводит её в выдачу.
    operator = _user_in_group(OPERATOR_GROUP)
    own = AppealFactory(created_by=operator, status=Appeal.Status.NEW)
    others = AppealFactory(status=Appeal.Status.NEW)
    client.login(email=operator.email, password="secret")

    # Запрос целится в чужой отдел: чужая заявка остаётся невидимой, а своя
    # заявка под условие чужого отдела закономерно не попадает.
    targeted = client.get(
        reverse(LIST_URL_NAME),
        {"status": Appeal.Status.NEW, "department": others.department_id},
    )
    assert others not in targeted.context["appeals"]
    assert own not in targeted.context["appeals"]

    # Без фильтра по отделу оператор по-прежнему видит только свою заявку.
    own_only = client.get(reverse(LIST_URL_NAME), {"status": Appeal.Status.NEW})
    appeals = list(own_only.context["appeals"])
    assert own in appeals
    assert others not in appeals


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_list_pagination_keeps_filters(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory.create_batch(25, created_by=operator, status=Appeal.Status.NEW)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(LIST_URL_NAME), {"status": Appeal.Status.NEW})

    assert response.context["is_paginated"] is True
    # Ссылки пагинации сохраняют активный фильтр (через querystring без page).
    assert response.context["querystring"] == f"status={Appeal.Status.NEW}"
    assert f"?status={Appeal.Status.NEW}&page=2" in response.content.decode()


# --- создание обращения: контроль доступа ------------------------------------


@pytest.mark.django_db
@pytest.mark.security
def test_appeal_create_redirects_anonymous_to_login(client):
    response = client.get(reverse(CREATE_URL_NAME))

    assert response.status_code == 302
    assert LOGIN_URL in response["Location"]


@pytest.mark.django_db
@pytest.mark.security
def test_appeal_create_forbidden_without_add_permission(client):
    # Ответственный сотрудник видит обращения, но не регистрирует новые.
    responsible = _user_in_group(RESPONSIBLE_GROUP)
    client.login(email=responsible.email, password="secret")

    response = client.get(reverse(CREATE_URL_NAME))

    assert response.status_code == 403


# --- создание обращения: отрисовка и отправка --------------------------------


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_create_renders_form_for_operator(client):
    operator = _user_in_group(OPERATOR_GROUP)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(CREATE_URL_NAME))

    assert response.status_code == 200
    assert "Новое обращение" in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_create_only_lists_active_categories(client):
    operator = _user_in_group(OPERATOR_GROUP)
    active = AppealCategoryFactory()
    inactive = AppealCategoryFactory(is_active=False)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(CREATE_URL_NAME))
    categories = list(response.context["form"].fields["category"].queryset)

    assert active in categories
    assert inactive not in categories


@pytest.mark.django_db
@pytest.mark.integration
def test_appeal_create_persists_appeal_and_redirects_to_detail(client):
    operator = _user_in_group(OPERATOR_GROUP)
    category = AppealCategoryFactory(default_processing_days=5)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(CREATE_URL_NAME),
        data=_valid_appeal_payload(category),
    )

    appeal = Appeal.objects.get()
    assert response.status_code == 302
    assert response["Location"] == reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk})
    assert appeal.created_by == operator
    assert appeal.department == category.department
    assert appeal.student_phone == "79001234567"
    assert appeal.status == Appeal.Status.NEW


@pytest.mark.django_db
@pytest.mark.integration
def test_appeal_create_writes_history_event(client):
    operator = _user_in_group(OPERATOR_GROUP)
    category = AppealCategoryFactory()
    client.login(email=operator.email, password="secret")

    client.post(reverse(CREATE_URL_NAME), data=_valid_appeal_payload(category))

    appeal = Appeal.objects.get()
    event = appeal.history_events.get()
    assert event.event_type == AppealHistoryEvent.EventType.CREATED
    assert event.actor == operator


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_create_shows_success_message(client):
    operator = _user_in_group(OPERATOR_GROUP)
    category = AppealCategoryFactory()
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(CREATE_URL_NAME),
        data=_valid_appeal_payload(category),
        follow=True,
    )

    assert "Обращение зарегистрировано." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_create_uses_selected_department(client):
    operator = _user_in_group(OPERATOR_GROUP)
    category = AppealCategoryFactory()
    other_department = DepartmentFactory()
    payload = _valid_appeal_payload(category)
    payload["department"] = other_department.pk
    client.login(email=operator.email, password="secret")

    client.post(reverse(CREATE_URL_NAME), data=payload)

    appeal = Appeal.objects.get()
    assert appeal.department == other_department


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_create_rejects_missing_fields(client):
    operator = _user_in_group(OPERATOR_GROUP)
    client.login(email=operator.email, password="secret")

    response = client.post(reverse(CREATE_URL_NAME), data={})

    assert response.status_code == 200
    assert Appeal.objects.count() == 0
    assert "Укажите ФИО студента." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.security
def test_appeal_create_rejects_inactive_category_choice(client):
    # Подделанный POST с неактивной категорией не является допустимым выбором.
    operator = _user_in_group(OPERATOR_GROUP)
    inactive = AppealCategoryFactory(is_active=False)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(CREATE_URL_NAME),
        data=_valid_appeal_payload(inactive),
    )

    assert response.status_code == 200
    assert Appeal.objects.count() == 0
    assert "Выберите категорию из списка." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_create_surfaces_service_validation_error(client):
    # Форма принимает активную категорию, но её отдел отключили после отрисовки —
    # сервис отклоняет заявку, и ошибка доходит до формы.
    operator = _user_in_group(OPERATOR_GROUP)
    department = DepartmentFactory(is_active=False)
    category = AppealCategoryFactory(department=department)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(CREATE_URL_NAME),
        data=_valid_appeal_payload(category),
    )

    assert response.status_code == 200
    assert Appeal.objects.count() == 0
    assert _("Selected department is inactive.") in response.content.decode()


# --- создание обращения с подбором категории через ИИ ------------------------


@pytest.fixture
def ai_on(settings):
    from core.ai import AIConfig

    settings.AI_CONFIG = AIConfig(
        model="deepseek-v4-flash",
        base_url="https://api.deepseek.com",
        api_key="sk-test",
    )


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_create_hides_classify_button_when_ai_disabled(client):
    # По умолчанию ИИ выключен: форма прежняя, кнопки подбора нет, поля видны.
    operator = _user_in_group(OPERATOR_GROUP)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(CREATE_URL_NAME))
    body = response.content.decode()

    assert "Подобрать категорию" not in body
    assert 'name="category"' in body


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_create_shows_classify_button_and_hides_route_fields(client, ai_on):
    # При включённом ИИ на первом этапе показываем подбор, а категорию/тему прячем.
    operator = _user_in_group(OPERATOR_GROUP)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(CREATE_URL_NAME))
    body = response.content.decode()

    assert "Подобрать категорию" in body
    assert 'name="category"' not in body


@pytest.mark.django_db
@pytest.mark.integration
def test_appeal_classify_shows_preview_and_prefills(client, ai_on):
    operator = _user_in_group(OPERATOR_GROUP)
    category = AppealCategoryFactory()
    client.login(email=operator.email, password="secret")
    result = ClassifyResult(
        status=ClassifyStatus.OK,
        category=category,
        summary="Подобранная тема",
        reason="Подходит по описанию.",
    )

    with patch("appeals.views.classify_appeal", return_value=result) as mocked:
        response = client.post(
            reverse(CREATE_URL_NAME),
            data={
                "action": "classify",
                "student_full_name": "Иванов Иван",
                "student_phone": "+7 900 123-45-67",
                "description": "Нужна справка об обучении",
            },
        )
    body = response.content.decode()

    mocked.assert_called_once_with("Нужна справка об обучении")
    assert response.status_code == 200
    assert Appeal.objects.count() == 0  # подбор ничего не сохраняет
    assert "Подходит по описанию." in body
    assert "Подобранная тема" in body
    assert 'name="category"' in body  # поля маршрута теперь видны для проверки
    assert "Подтвердить и сохранить" in body


@pytest.mark.django_db
@pytest.mark.integration
def test_appeal_classify_undecided_shows_notice(client, ai_on):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealCategoryFactory()
    client.login(email=operator.email, password="secret")
    result = ClassifyResult(status=ClassifyStatus.UNDECIDED, reason="Не относится ни к чему.")

    with patch("appeals.views.classify_appeal", return_value=result):
        response = client.post(
            reverse(CREATE_URL_NAME),
            data={
                "action": "classify",
                "student_full_name": "Иванов Иван",
                "student_phone": "+7 900 123-45-67",
                "description": "во сколько обед",
            },
        )
    body = response.content.decode()

    assert "Не удалось определить категорию" in body
    assert 'name="category"' in body  # можно выбрать вручную


@pytest.mark.django_db
@pytest.mark.integration
def test_appeal_classify_unavailable_shows_calm_notice(client, ai_on):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealCategoryFactory()
    client.login(email=operator.email, password="secret")
    result = ClassifyResult(status=ClassifyStatus.UNAVAILABLE)

    with patch("appeals.views.classify_appeal", return_value=result):
        response = client.post(
            reverse(CREATE_URL_NAME),
            data={
                "action": "classify",
                "student_full_name": "Иванов Иван",
                "student_phone": "+7 900 123-45-67",
                "description": "нужна справка",
            },
        )
    body = response.content.decode()

    assert "ИИ-подбор сейчас недоступен" in body
    assert 'name="category"' in body


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_classify_requires_description(client, ai_on):
    operator = _user_in_group(OPERATOR_GROUP)
    client.login(email=operator.email, password="secret")

    with patch("appeals.views.classify_appeal") as mocked:
        response = client.post(
            reverse(CREATE_URL_NAME),
            data={"action": "classify", "student_full_name": "Иванов", "description": ""},
        )

    mocked.assert_not_called()  # без описания модель не дёргаем
    assert response.status_code == 200
    assert "Опишите суть обращения для подбора категории." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.integration
def test_appeal_save_after_classify_persists(client, ai_on):
    # После подбора оператор подтверждает — обращение сохраняется обычным путём.
    operator = _user_in_group(OPERATOR_GROUP)
    category = AppealCategoryFactory()
    client.login(email=operator.email, password="secret")

    payload = _valid_appeal_payload(category)
    payload["action"] = "save"
    response = client.post(reverse(CREATE_URL_NAME), data=payload)

    assert response.status_code == 302
    appeal = Appeal.objects.get()
    assert appeal.category == category
    assert appeal.created_by == operator


# --- карточка обращения ------------------------------------------------------


@pytest.mark.django_db
@pytest.mark.security
def test_appeal_detail_redirects_anonymous_to_login(client):
    appeal = AppealFactory()
    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 302
    assert LOGIN_URL in response["Location"]


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_detail_renders_for_owner(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator, summary="моя заявка")
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 200
    assert "моя заявка" in response.content.decode()
    assert appeal.student_full_name in response.content.decode()


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_appeal_detail_hidden_appeal_returns_404(client):
    operator = _user_in_group(OPERATOR_GROUP)
    other = AppealFactory(summary="чужая заявка")
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": other.pk}))

    assert response.status_code == 404


@pytest.mark.django_db
@pytest.mark.security
def test_appeal_detail_forbidden_without_view_permission(client):
    appeal = AppealFactory()
    user = UserFactory(password="secret")
    client.login(email=user.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 403


@pytest.mark.django_db
@pytest.mark.security
def test_appeal_detail_denies_deactivated_user_with_live_session(client):
    # Отключение пользователя обрывает его активную сессию: бэкенд аутентификации
    # перестаёт сопоставлять сессию с пользователем, поэтому следующий запрос
    # становится анонимным и уходит на вход, а не показывает обращение.
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator, summary="секретная заявка")
    client.login(email=operator.email, password="secret")
    operator.is_active = False
    operator.save(update_fields=["is_active"])

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 302
    assert LOGIN_URL in response["Location"]
    assert "секретная заявка" not in response.content.decode()


# --- карточка обращения: отрисовка истории и комментариев ---------------------


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_detail_shows_history_timeline(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator)
    AppealHistoryEventFactory(
        appeal=appeal,
        actor=operator,
        event_type=AppealHistoryEvent.EventType.CREATED,
    )
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))
    content = response.content.decode()

    assert list(response.context["history_events"]) == list(appeal.history_events.all())
    assert "История" in content
    assert "timeline" in content


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_detail_lists_comments(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator)
    AppealCommentFactory(appeal=appeal, author=operator, text="комментарий оператора")
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert "комментарий оператора" in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_detail_shows_comment_form_for_commenter(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.context["can_comment"] is True
    assert "Добавить комментарий" in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_appeal_detail_lets_responsible_comment_department_appeal(client):
    responsible = _user_in_group(RESPONSIBLE_GROUP)
    department = DepartmentFactory()
    department.members.add(responsible)
    appeal = AppealFactory(department=department)
    client.login(email=responsible.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.context["can_comment"] is True
    assert "Добавить комментарий" in response.content.decode()


# --- добавление комментария --------------------------------------------------


@pytest.mark.django_db
@pytest.mark.security
def test_comment_create_redirects_anonymous_to_login(client):
    appeal = AppealFactory()
    response = client.post(reverse(COMMENT_URL_NAME, kwargs={"pk": appeal.pk}), data={"text": "x"})

    assert response.status_code == 302
    assert LOGIN_URL in response["Location"]


@pytest.mark.django_db
@pytest.mark.security
def test_comment_create_forbidden_without_comment_permission(client):
    appeal = AppealFactory()
    user = UserFactory(password="secret")
    client.login(email=user.email, password="secret")

    response = client.post(
        reverse(COMMENT_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"text": "Комментарий"},
    )

    assert response.status_code == 403
    assert AppealComment.objects.count() == 0


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_comment_create_returns_404_for_out_of_scope_appeal(client):
    # У оператора есть право комментировать в целом, но чужие обращения он не
    # видит, поэтому комментирование такого — это 404 (без раскрытия), а не 403.
    operator = _user_in_group(OPERATOR_GROUP)
    other = AppealFactory()
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(COMMENT_URL_NAME, kwargs={"pk": other.pk}),
        data={"text": "Комментарий"},
    )

    assert response.status_code == 404
    assert AppealComment.objects.count() == 0


@pytest.mark.django_db
@pytest.mark.functional
def test_comment_create_rejects_get(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(COMMENT_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 405


@pytest.mark.django_db
@pytest.mark.integration
def test_comment_create_adds_comment_and_redirects(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(COMMENT_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"text": "Уточнил детали обращения"},
    )

    comment = AppealComment.objects.get()
    assert comment.appeal == appeal
    assert comment.author == operator
    assert comment.text == "Уточнил детали обращения"
    assert response.status_code == 302
    detail_url = reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk})
    assert response["Location"] == f"{detail_url}#comments"


@pytest.mark.django_db
@pytest.mark.integration
def test_comment_create_writes_history_event(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    client.post(
        reverse(COMMENT_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"text": "Комментарий"},
    )

    event = appeal.history_events.get()
    assert event.event_type == AppealHistoryEvent.EventType.COMMENT_ADDED
    assert event.actor == operator


@pytest.mark.django_db
@pytest.mark.functional
def test_comment_create_shows_success_message(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(COMMENT_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"text": "Комментарий"},
        follow=True,
    )

    assert "Комментарий добавлен." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_comment_create_rejects_empty_text(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(COMMENT_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"text": "   "},
        follow=True,
    )

    assert AppealComment.objects.count() == 0
    assert "Введите текст комментария." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.integration
def test_comment_create_rejects_closed_appeal(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator, status=Appeal.Status.CLOSED)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(COMMENT_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"text": "Комментарий к закрытой"},
        follow=True,
    )

    assert AppealComment.objects.count() == 0
    assert "Нельзя комментировать закрытое обращение." in response.content.decode()


# --- взятие обращения в работу -----------------------------------------------


@pytest.mark.django_db
@pytest.mark.security
def test_start_processing_redirects_anonymous_to_login(client):
    appeal = AppealFactory()
    response = client.post(reverse(START_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 302
    assert LOGIN_URL in response["Location"]


@pytest.mark.django_db
@pytest.mark.security
def test_start_processing_forbidden_without_permission(client):
    # Оператор видит свои обращения, но не может брать их в работу.
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator, status=Appeal.Status.NEW)
    client.login(email=operator.email, password="secret")

    response = client.post(reverse(START_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 403
    appeal.refresh_from_db()
    assert appeal.status == Appeal.Status.NEW


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_start_processing_returns_404_for_out_of_scope_appeal(client):
    responsible, _ = _responsible_with_department()
    other = AppealFactory(status=Appeal.Status.NEW)
    client.login(email=responsible.email, password="secret")

    response = client.post(reverse(START_URL_NAME, kwargs={"pk": other.pk}))

    assert response.status_code == 404
    other.refresh_from_db()
    assert other.status == Appeal.Status.NEW


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_start_processing_forbidden_for_visible_appeal_outside_department(client):
    # Пользователь сразу в ролях оператора и ответственного видит созданное им
    # обращение, даже если оно в отделе, где он не состоит. Право брать в работу
    # у него есть, поэтому запрос проходит миксин и выборку по доступу, но
    # проверка по отделу для конкретного объекта всё равно запрещает действие.
    user = _user_in_group(OPERATOR_GROUP, RESPONSIBLE_GROUP)
    other_department = DepartmentFactory()
    appeal = AppealFactory(
        created_by=user,
        department=other_department,
        status=Appeal.Status.NEW,
    )
    client.login(email=user.email, password="secret")

    response = client.post(reverse(START_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 403
    appeal.refresh_from_db()
    assert appeal.status == Appeal.Status.NEW


@pytest.mark.django_db
@pytest.mark.functional
def test_start_processing_rejects_get(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department, status=Appeal.Status.NEW)
    client.login(email=responsible.email, password="secret")

    response = client.get(reverse(START_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 405


@pytest.mark.django_db
@pytest.mark.integration
def test_start_processing_moves_appeal_into_work(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department, status=Appeal.Status.NEW)
    client.login(email=responsible.email, password="secret")

    response = client.post(reverse(START_URL_NAME, kwargs={"pk": appeal.pk}))

    appeal.refresh_from_db()
    assert appeal.status == Appeal.Status.IN_PROGRESS
    assert appeal.accepted_by == responsible
    assert response.status_code == 302
    assert response["Location"] == reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk})

    event = appeal.history_events.latest("created_at")
    assert event.event_type == AppealHistoryEvent.EventType.ACCEPTED
    assert event.actor == responsible


@pytest.mark.django_db
@pytest.mark.functional
def test_start_processing_shows_success_message(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department, status=Appeal.Status.NEW)
    client.login(email=responsible.email, password="secret")

    response = client.post(
        reverse(START_URL_NAME, kwargs={"pk": appeal.pk}),
        follow=True,
    )

    assert "Обращение взято в работу." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.integration
def test_start_processing_rejects_appeal_already_in_work(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department, status=Appeal.Status.IN_PROGRESS)
    client.login(email=responsible.email, password="secret")

    response = client.post(
        reverse(START_URL_NAME, kwargs={"pk": appeal.pk}),
        follow=True,
    )

    assert "Обращение уже взято в работу." in response.content.decode()


# --- закрытие обращения ------------------------------------------------------


@pytest.mark.django_db
@pytest.mark.security
def test_close_redirects_anonymous_to_login(client):
    appeal = AppealFactory()
    response = client.get(reverse(CLOSE_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 302
    assert LOGIN_URL in response["Location"]


@pytest.mark.django_db
@pytest.mark.security
def test_close_forbidden_without_permission(client):
    # Ответственный без этого обращения в своём отделе не может его закрыть, а
    # пользователя совсем без права закрытия отклоняет миксин.
    user = UserFactory(password="secret")
    appeal = AppealFactory()
    client.login(email=user.email, password="secret")

    response = client.get(reverse(CLOSE_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 403


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_close_returns_404_for_out_of_scope_appeal(client):
    operator = _user_in_group(OPERATOR_GROUP)
    other = AppealFactory()
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(CLOSE_URL_NAME, kwargs={"pk": other.pk}))

    assert response.status_code == 404


@pytest.mark.django_db
@pytest.mark.functional
def test_close_renders_form_for_owner(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator, status=Appeal.Status.IN_PROGRESS)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(CLOSE_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 200
    assert "Результат обработки" in response.content.decode()


@pytest.mark.django_db
@pytest.mark.integration
def test_close_sets_result_and_status(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator, status=Appeal.Status.IN_PROGRESS)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(CLOSE_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"result": "Справка выдана студенту."},
    )

    appeal.refresh_from_db()
    assert appeal.status == Appeal.Status.CLOSED
    assert appeal.result == "Справка выдана студенту."
    assert response.status_code == 302
    assert response["Location"] == reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk})

    event = appeal.history_events.latest("created_at")
    assert event.event_type == AppealHistoryEvent.EventType.CLOSED
    assert event.actor == operator


@pytest.mark.django_db
@pytest.mark.functional
def test_close_shows_success_message(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator, status=Appeal.Status.IN_PROGRESS)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(CLOSE_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"result": "Готово."},
        follow=True,
    )

    assert "Обращение закрыто." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_close_rejects_empty_result(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator, status=Appeal.Status.IN_PROGRESS)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(CLOSE_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"result": "   "},
    )

    assert response.status_code == 200
    appeal.refresh_from_db()
    assert appeal.status == Appeal.Status.IN_PROGRESS
    assert "Опишите результат обработки." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.integration
def test_close_already_closed_appeal_surfaces_service_error(client):
    # После закрытия кнопка закрытия скрыта, но прямой POST всё равно
    # отклоняется сервисом и показывается как ошибка формы.
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator, status=Appeal.Status.CLOSED)
    client.login(email=operator.email, password="secret")

    response = client.post(
        reverse(CLOSE_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"result": "Повторное закрытие."},
    )

    assert response.status_code == 200
    assert _("Appeal is already closed.") in response.content.decode()


# --- карточка обращения: доступность действий ---------------------------------


@pytest.mark.django_db
@pytest.mark.functional
def test_detail_shows_take_into_work_for_responsible_on_new(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department, status=Appeal.Status.NEW)
    client.login(email=responsible.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.context["can_start_processing"] is True
    assert "Взять в работу" in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_detail_hides_take_into_work_once_in_progress(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department, status=Appeal.Status.IN_PROGRESS)
    client.login(email=responsible.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.context["can_start_processing"] is False


@pytest.mark.django_db
@pytest.mark.functional
def test_detail_shows_close_button_for_owner_until_closed(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator, status=Appeal.Status.IN_PROGRESS)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.context["can_close"] is True
    assert "Закрыть обращение" in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_detail_hides_actions_and_shows_result_when_closed(client):
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(
        created_by=operator,
        status=Appeal.Status.CLOSED,
        result="Итоговый результат обработки.",
    )
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))
    content = response.content.decode()

    assert response.context["can_close"] is False
    assert response.context["can_start_processing"] is False
    assert "Итоговый результат обработки." in content


# --- перенаправление обращения -----------------------------------------------


@pytest.mark.django_db
@pytest.mark.security
def test_transfer_redirects_anonymous_to_login(client):
    appeal = AppealFactory()
    response = client.get(reverse(TRANSFER_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 302
    assert LOGIN_URL in response["Location"]


@pytest.mark.django_db
@pytest.mark.security
def test_transfer_forbidden_without_permission(client):
    # Оператор может видеть свои обращения, но не может их перенаправлять.
    operator = _user_in_group(OPERATOR_GROUP)
    appeal = AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(TRANSFER_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 403


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_transfer_returns_404_for_out_of_scope_appeal(client):
    responsible, _ = _responsible_with_department()
    other = AppealFactory()
    client.login(email=responsible.email, password="secret")

    response = client.get(reverse(TRANSFER_URL_NAME, kwargs={"pk": other.pk}))

    assert response.status_code == 404


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_transfer_forbidden_for_visible_appeal_outside_department(client):
    # Двойная роль: своё созданное обращение видно даже в чужом отделе, право на
    # перенос есть, но проверка по отделу для объекта запрещает действие.
    user = _user_in_group(OPERATOR_GROUP, RESPONSIBLE_GROUP)
    other_department = DepartmentFactory()
    appeal = AppealFactory(created_by=user, department=other_department)
    client.login(email=user.email, password="secret")

    response = client.get(reverse(TRANSFER_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 403


@pytest.mark.django_db
@pytest.mark.functional
def test_transfer_form_is_prefilled_with_current_route(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department)
    client.login(email=responsible.email, password="secret")

    response = client.get(reverse(TRANSFER_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.status_code == 200
    initial = response.context["form"].initial
    assert initial["category"] == appeal.category_id
    assert initial["department"] == appeal.department_id


@pytest.mark.django_db
@pytest.mark.integration
def test_transfer_changes_category_and_writes_history(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department)
    new_category = AppealCategoryFactory(department=department)
    client.login(email=responsible.email, password="secret")

    response = client.post(
        reverse(TRANSFER_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"category": new_category.pk, "department": department.pk},
    )

    appeal.refresh_from_db()
    assert appeal.category == new_category
    assert appeal.department == department
    assert response.status_code == 302
    assert response["Location"] == reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk})

    event = appeal.history_events.latest("created_at")
    assert event.event_type == AppealHistoryEvent.EventType.CATEGORY_CHANGED
    assert event.actor == responsible


@pytest.mark.django_db
@pytest.mark.integration
def test_transfer_changes_department(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department)
    target_department = DepartmentFactory()
    target_department.members.add(responsible)
    client.login(email=responsible.email, password="secret")

    client.post(
        reverse(TRANSFER_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"category": appeal.category_id, "department": target_department.pk},
    )

    appeal.refresh_from_db()
    assert appeal.department == target_department

    event = appeal.history_events.latest("created_at")
    assert event.event_type == AppealHistoryEvent.EventType.DEPARTMENT_CHANGED


@pytest.mark.django_db
@pytest.mark.integration
def test_transfer_to_other_department_redirects_to_list(client):
    # Перенос в чужой отдел выводит заявку из зоны видимости сотрудника, поэтому
    # после успеха возвращаемся к списку, а не в недоступную карточку (404).
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department)
    other_department = DepartmentFactory()  # ответственный в нём не состоит
    target_category = AppealCategoryFactory(department=other_department)
    client.login(email=responsible.email, password="secret")

    response = client.post(
        reverse(TRANSFER_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"category": target_category.pk, "department": other_department.pk},
    )

    appeal.refresh_from_db()
    assert appeal.department == other_department
    assert response.status_code == 302
    assert response["Location"] == reverse(LIST_URL_NAME)


@pytest.mark.django_db
@pytest.mark.functional
def test_transfer_shows_success_message(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department)
    new_category = AppealCategoryFactory(department=department)
    client.login(email=responsible.email, password="secret")

    response = client.post(
        reverse(TRANSFER_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"category": new_category.pk, "department": department.pk},
        follow=True,
    )

    assert "Обращение перенаправлено." in response.content.decode()


@pytest.mark.django_db
@pytest.mark.integration
def test_transfer_rejects_unchanged_route(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department)
    client.login(email=responsible.email, password="secret")

    response = client.post(
        reverse(TRANSFER_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"category": appeal.category_id, "department": appeal.department_id},
    )

    assert response.status_code == 200
    assert _("Appeal route is unchanged.") in response.content.decode()


@pytest.mark.django_db
@pytest.mark.integration
def test_transfer_rejects_closed_appeal(client):
    # После закрытия кнопка переноса скрыта, но прямой POST отклоняется сервисом.
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department, status=Appeal.Status.CLOSED)
    new_category = AppealCategoryFactory(department=department)
    client.login(email=responsible.email, password="secret")

    response = client.post(
        reverse(TRANSFER_URL_NAME, kwargs={"pk": appeal.pk}),
        data={"category": new_category.pk, "department": department.pk},
    )

    assert response.status_code == 200
    appeal.refresh_from_db()
    assert appeal.category != new_category


@pytest.mark.django_db
@pytest.mark.functional
def test_detail_shows_transfer_button_for_responsible(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department, status=Appeal.Status.NEW)
    client.login(email=responsible.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.context["can_transfer"] is True
    assert "Перенаправить" in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_detail_hides_transfer_button_when_closed(client):
    responsible, department = _responsible_with_department()
    appeal = AppealFactory(department=department, status=Appeal.Status.CLOSED)
    client.login(email=responsible.email, password="secret")

    response = client.get(reverse(DETAIL_URL_NAME, kwargs={"pk": appeal.pk}))

    assert response.context["can_transfer"] is False


# --- отчёты: доступ ----------------------------------------------------------


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.parametrize(
    "url_name",
    [REPORT_URL_NAME, REPORT_XLSX_URL_NAME, REPORT_DOCX_URL_NAME],
)
def test_report_redirects_anonymous_to_login(client, url_name):
    response = client.get(reverse(url_name))

    assert response.status_code == 302
    assert LOGIN_URL in response["Location"]


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.parametrize(
    "url_name",
    [REPORT_URL_NAME, REPORT_XLSX_URL_NAME, REPORT_DOCX_URL_NAME],
)
def test_report_forbidden_without_view_permission(client, url_name):
    user = UserFactory(password="secret")
    client.login(email=user.email, password="secret")

    response = client.get(reverse(url_name))

    assert response.status_code == 403


# --- отчёты: содержимое страницы и охват -------------------------------------


@pytest.mark.django_db
@pytest.mark.functional
def test_report_page_renders_summary(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory(created_by=operator, status=Appeal.Status.NEW)
    AppealFactory(created_by=operator, status=Appeal.Status.CLOSED)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(REPORT_URL_NAME))

    assert response.status_code == 200
    assert response.context["report"].total == 2
    content = response.content.decode()
    assert "Отчёты по обращениям" in content
    assert "Скачать .xlsx" in content
    assert "Скачать .docx" in content


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_report_scopes_to_operator_own_appeals(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory(created_by=operator)
    AppealFactory()  # чужая заявка не должна попасть в сводку

    client.login(email=operator.email, password="secret")
    response = client.get(reverse(REPORT_URL_NAME))

    assert response.context["report"].total == 1


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_report_scopes_to_responsible_department(client):
    responsible, department = _responsible_with_department()
    AppealFactory(department=department)
    AppealFactory()  # заявка чужого отдела

    client.login(email=responsible.email, password="secret")
    response = client.get(reverse(REPORT_URL_NAME))

    assert response.context["report"].total == 1


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_report_covers_all_appeals_for_admin(client):
    admin = _user_in_group(ADMIN_GROUP)
    AppealFactory()
    AppealFactory()

    client.login(email=admin.email, password="secret")
    response = client.get(reverse(REPORT_URL_NAME))

    assert response.context["report"].total == 2


# --- отчёты: выгрузка файлов -------------------------------------------------


@pytest.mark.django_db
@pytest.mark.functional
def test_report_xlsx_download_headers(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(REPORT_XLSX_URL_NAME))

    assert response.status_code == 200
    assert response["Content-Type"] == XLSX_CONTENT_TYPE
    assert "attachment" in response["Content-Disposition"]
    assert "appeal-report.xlsx" in response["Content-Disposition"]
    assert response.content


@pytest.mark.django_db
@pytest.mark.functional
def test_report_docx_download_headers(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(REPORT_DOCX_URL_NAME))

    assert response.status_code == 200
    assert response["Content-Type"] == DOCX_CONTENT_TYPE
    assert "attachment" in response["Content-Disposition"]
    assert "appeal-report.docx" in response["Content-Disposition"]
    assert response.content


@pytest.mark.django_db
@pytest.mark.security
@pytest.mark.integration
def test_report_xlsx_download_respects_scope(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory(created_by=operator)
    AppealFactory()  # чужая заявка
    client.login(email=operator.email, password="secret")

    response = client.get(reverse(REPORT_XLSX_URL_NAME))

    workbook = load_workbook(BytesIO(response.content))
    summary = workbook["Сводка"]
    totals = {
        summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value
        for row in range(1, summary.max_row + 1)
    }
    assert totals["Всего обращений"] == 1


# --- отчёты: сохранённые на диск файлы ---------------------------------------


@pytest.mark.django_db
@pytest.mark.integration
def test_report_export_saves_file_to_disk(client, settings):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    client.get(reverse(REPORT_XLSX_URL_NAME))

    # После выгрузки в каталоге reports появляется сохранённый файл.
    reports_dir = settings.MEDIA_ROOT / "reports"
    saved = list(reports_dir.glob("*.xlsx"))
    assert len(saved) == 1


@pytest.mark.django_db
@pytest.mark.functional
def test_report_page_lists_saved_files(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    # Сначала формируем файл, затем он должен появиться в списке на странице.
    client.get(reverse(REPORT_DOCX_URL_NAME))
    response = client.get(reverse(REPORT_URL_NAME))

    stored = response.context["stored_reports"]
    assert len(stored) == 1
    assert stored[0].name in response.content.decode()


@pytest.mark.django_db
@pytest.mark.functional
def test_saved_report_file_downloads_with_content(client):
    operator = _user_in_group(OPERATOR_GROUP)
    AppealFactory(created_by=operator)
    client.login(email=operator.email, password="secret")

    exported = client.get(reverse(REPORT_XLSX_URL_NAME))
    page = client.get(reverse(REPORT_URL_NAME))
    name = page.context["stored_reports"][0].name

    response = client.get(reverse(REPORT_FILE_URL_NAME, kwargs={"name": name}))

    assert response.status_code == 200
    assert response["Content-Type"] == XLSX_CONTENT_TYPE
    assert name in response["Content-Disposition"]
    # С диска отдаётся ровно то, что было выгружено.
    assert response.content == exported.content


@pytest.mark.django_db
@pytest.mark.functional
def test_saved_report_file_missing_returns_404(client):
    operator = _user_in_group(OPERATOR_GROUP)
    client.login(email=operator.email, password="secret")

    response = client.get(
        reverse(REPORT_FILE_URL_NAME, kwargs={"name": "appeal-report-00000000-000000.xlsx"}),
    )

    assert response.status_code == 404


@pytest.mark.django_db
@pytest.mark.security
def test_saved_report_file_redirects_anonymous_to_login(client):
    response = client.get(
        reverse(REPORT_FILE_URL_NAME, kwargs={"name": "appeal-report.xlsx"}),
    )

    assert response.status_code == 302
    assert LOGIN_URL in response["Location"]


@pytest.mark.django_db
@pytest.mark.security
def test_saved_report_file_forbidden_without_view_permission(client):
    user = UserFactory(password="secret")
    client.login(email=user.email, password="secret")

    response = client.get(
        reverse(REPORT_FILE_URL_NAME, kwargs={"name": "appeal-report.xlsx"}),
    )

    assert response.status_code == 403
